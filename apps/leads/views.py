from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Count, Prefetch
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse
import json
from django.utils import timezone
from datetime import datetime
from django.views.decorators.http import require_POST
from apps.accounts.decorators import company_required, admin_required
from .models import Lead, Note, Activity
from apps.core.models import LeadSource, LeadStage
from .forms import *
from django.http import JsonResponse, HttpResponse
from apps.accounts.models import User
import openpyxl
from openpyxl.styles import Font, PatternFill
import csv
from io import StringIO, TextIOWrapper


@login_required
@company_required
def lead_list_view(request):
    company = request.user.company
    leads = Lead.objects.filter(company=company).exclude(status='deleted').select_related('source','stage','assigned_to').order_by('-created_at')

    search_query = request.GET.get('search', '').strip()

    if search_query:
        # Search in name, phone, or email using Q objects (OR condition)
        leads = leads.filter(
            Q(name__icontains=search_query) |  # Name contains search term
            Q(phone__icontains=search_query) |  # Phone contains search term
            Q(email__icontains=search_query)  # Email contains search term
        )

    filter_form = LeadFilterForm(request.GET, company=company)

    if filter_form.is_valid():
        # Filter by source
        if filter_form.cleaned_data.get('source'):
            leads = leads.filter(source=filter_form.cleaned_data['source'])

        # Filter by stage
        if filter_form.cleaned_data.get('stage'):
            leads = leads.filter(stage=filter_form.cleaned_data['stage'])

        # Filter by status
        if filter_form.cleaned_data.get('status'):
            leads = leads.filter(status=filter_form.cleaned_data['status'])

        # Filter by priority
        if filter_form.cleaned_data.get('priority'):
            leads = leads.filter(priority=filter_form.cleaned_data['priority'])

        # Filter by assigned agent
        if filter_form.cleaned_data.get('assigned_to'):
            leads = leads.filter(assigned_to=filter_form.cleaned_data['assigned_to'])

        # Filter by date range (created_at)
        if filter_form.cleaned_data.get('date_from'):
            # Created on or after this date
            leads = leads.filter(created_at__date__gte=filter_form.cleaned_data['date_from'])

        if filter_form.cleaned_data.get('date_to'):
            # Created on or before this date
            leads = leads.filter(created_at__date__lte=filter_form.cleaned_data['date_to'])

    total_count = leads.count()
    status_counts = {
        'new': leads.filter(status='new').count(),
        'contacted': leads.filter(status='contacted').count(),
        'qualified': leads.filter(status='qualified').count(),
        'converted': leads.filter(status='converted').count(),
        'won': leads.filter(status='won').count(),
        'lost': leads.filter(status='lost').count(),
    }

    paginator = Paginator(leads, 50)
    page_number = request.GET.get('page', 1)

    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        # If page is not an integer, deliver first page
        page_obj = paginator.get_page(1)
    except EmptyPage:
        # If page is out of range, deliver last page
        page_obj = paginator.get_page(paginator.num_pages)


    context = {
        'leads': page_obj,
        'filter_form': filter_form,
        'total_count': total_count,
        'status_counts': status_counts,
        'page_obj': page_obj,
        'search_query': search_query,

        # Pagination info
        'is_paginated': page_obj.has_other_pages(),
        'page_range': paginator.get_elided_page_range(
            page_obj.number,
            on_each_side=2,
            on_ends=1
        ),
        
        # Modal Data
        'agents': User.objects.filter(company=company, is_active=True).order_by('first_name'),
        'stages': LeadStage.objects.filter(is_active=True).order_by('order'),
    }

    return render(request, 'leads/lead_list.html', context)


@login_required
@company_required
def lead_detail_view(request, pk):
    lead = get_object_or_404(
        Lead.objects.select_related(
            'company',
            'source',
            'stage',
            'assigned_to',
            'assigned_to__profile'
        ),
        pk=pk,
        company=request.user.company
    )

    if request.method == 'POST':
        note_form = NoteForm(request.POST)

        if note_form.is_valid():
            note = lead.add_note(
                content=note_form.cleaned_data['content'],
                user=request.user
            )
            messages.success(request, 'Note added successfully')

            return redirect('leads:lead_detail', pk=lead.pk)
        else:
            messages.error(request, 'Error adding note')
    else:
        note_form = NoteForm()


    notes = lead.get_notes()

    activities = lead.get_activities()


    # User must be admin OR assigned to this lead
    can_edit = (
            request.user.is_admin() or
            lead.assigned_to == request.user
    )

    # Only admins can delete
    can_delete = request.user.is_admin()


    # Rules: Lead must not be Won/Lost
    can_be_assigned = lead.can_be_assigned()

    context = {
        'lead': lead,
        'notes': notes,
        'activities': activities,
        'note_form': note_form,
        'can_edit': can_edit,
        'can_delete': can_delete,
        'can_be_assigned': can_be_assigned,

        # Counts for tabs
        'notes_count': notes.count(),
        'activities_count': activities.count(),

        # Current tab (from query parameter)
        'active_tab': request.GET.get('tab', 'overview'),
        
        # Modal Data
        'agents': User.objects.filter(company=request.user.company, is_active=True).order_by('first_name'),
        'stages': LeadStage.objects.filter(is_active=True).order_by('order'),
    }

    return render(request, 'leads/lead_detail.html', context)


@login_required
@company_required
def lead_json_view(request, pk):

    try:
        lead = Lead.objects.select_related(
            'source',
            'stage',
            'assigned_to'
        ).get(
            pk=pk,
            company=request.user.company
        )

        data = {
            'id': lead.id,
            'name': lead.name,
            'phone': lead.phone,
            'email': lead.email,

            'source': {
                'id': lead.source.id,
                'name': lead.source.name,
                'color': lead.source.color,
                'icon': lead.source.icon,
            },

            # Stage info
            'stage': {
                'id': lead.stage.id,
                'name': lead.stage.name,
                'color': lead.stage.color,
                'icon': lead.stage.icon,
            },

            # Status
            'status': lead.status,
            'status_display': lead.get_status_display(),

            # Priority
            'priority': lead.priority,
            'priority_display': lead.get_priority_display(),

            # Assignment
            'assigned_to': {
                'id': lead.assigned_to.id,
                'name': lead.assigned_to.get_full_name(),
            } if lead.assigned_to else None,

            # Dates
            'next_follow_up': lead.next_follow_up.isoformat() if lead.next_follow_up else None,
            'created_at': lead.created_at.isoformat(),
            'updated_at': lead.updated_at.isoformat(),

            # Additional info
            'notes': lead.notes,
            'tags': lead.tags,

            # Computed fields
            'time_since_created': lead.time_since_created(),
            'time_until_follow_up': lead.time_until_follow_up(),
            'can_be_assigned': lead.can_be_assigned(),
        }

        return JsonResponse(data)

    except Lead.DoesNotExist:
        return JsonResponse(
            {'error': 'Lead not found'},
            status=404
        )
    except Exception as e:
        return JsonResponse(
            {'error': str(e)},
            status=500
        )


@login_required
@company_required
def lead_activities_view(request, pk):
    try:
        lead = get_object_or_404(
            Lead,
            pk=pk,
            company=request.user.company
        )

        activities = Activity.objects.filter(
            lead=lead
        ).select_related('user').order_by('-created_at')

        page_number = request.GET.get('page', 1)
        per_page = int(request.GET.get('per_page', 20))

        paginator = Paginator(activities, per_page)
        page_obj = paginator.get_page(page_number)

        # Build response
        activities_data = []
        for activity in page_obj:
            activities_data.append({
                'id': activity.id,
                'type': activity.activity_type,
                'type_display': activity.get_activity_type_display(),
                'description': activity.description,
                'user': {
                    'id': activity.user.id,
                    'name': activity.user.get_full_name(),
                    'initials': activity.user.get_initials(),
                } if activity.user else None,
                'created_at': activity.created_at.isoformat(),
                'time_since': activity.created_at,  # Will be formatted in frontend
            })

        data = {
            'activities': activities_data,
            'has_next': page_obj.has_next(),
            'has_previous': page_obj.has_previous(),
            'total_count': paginator.count,
            'page': page_obj.number,
            'num_pages': paginator.num_pages,
        }

        return JsonResponse(data)

    except Lead.DoesNotExist:
        return JsonResponse(
            {'error': 'Lead not found'},
            status=404
        )
    except Exception as e:
        return JsonResponse(
            {'error': str(e)},
            status=500
        )


@login_required
@company_required
def lead_create_view(request):
    if request.method == 'POST':
        # Bind the form with POST data and pass the user's company
        form = LeadCreateForm(request.POST, company=request.user.company)

        if form.is_valid():
            try:
                lead = form.save(commit=False)
                lead.company = request.user.company

                if not lead.assigned_to:
                    lead.assigned_to = request.user
                lead.save()

                # Save many-to-many relationships (if any)
                form.save_m2m()

                Activity.objects.create(
                    lead=lead,
                    user=request.user,
                    activity_type='created',
                    description=f'Lead created by {request.user.get_full_name()}'
                )


                if lead.assigned_to:
                    lead.assigned_to.total_leads_assigned += 1
                    lead.assigned_to.save(update_fields=['total_leads_assigned'])

                messages.success(
                    request,
                    f'Lead "{lead.name}" created successfully'
                )
                return redirect('leads:lead_detail', pk=lead.pk)

            except Exception as e:
                # Handle any unexpected errors during lead creation
                messages.error(request, f'Error creating lead: {str(e)}')

        # If the form is invalid or an exception occurred, re-render the form
        context = {
            'form': form,
            'form_title': 'Create New Lead',
            'submit_text': 'Create',
            'cancel_url': 'leads:lead_list',
        }
        return render(request, 'leads/lead_form.html', context)

    else:
        # GET request: render an empty form for creating a lead
        form = LeadCreateForm(company=request.user.company)
        context = {
            'form': form,
            'form_title': 'Create New Lead',
            'submit_text': 'Create',
            'cancel_url': 'leads:lead_list',
        }
        return render(request, 'leads/lead_form.html', context)


@company_required
def lead_edit_view(request, pk):
    lead = get_object_or_404(
        Lead,
        pk=pk,
        company=request.user.company
    )

    if not (request.user.is_admin() or lead.assigned_to == request.user):
        messages.error(request, 'You do not have permission to edit this lead')
        return redirect('leads:lead_detail', pk=lead.pk)

    if request.method == 'POST':
        form = LeadEditForm(
            request.POST,
            instance=lead,
            company=request.user.company
        )

        if form.is_valid():
            try:
                changed_fields = []

                for field in form.changed_data:
                    old_value = getattr(lead, field)
                    new_value = form.cleaned_data[field]

                    if old_value != new_value:
                        # Get human-readable field name
                        field_label = form.fields[field].label or field
                        changed_fields.append(field_label)

                lead = form.save()

                if changed_fields:
                    Activity.objects.create(
                        lead=lead,
                        user=request.user,
                        activity_type='status_changed',
                        description=f'Updated: {", ".join(changed_fields)}'
                    )

                messages.success(
                    request,
                    f'Lead "{lead.name}" updated successfully'
                )

                return redirect('leads:lead_detail', pk=lead.pk)

            except Exception as e:
                messages.error(
                    request,
                    f'Error updating lead: {str(e)}'
                )
        else:
            messages.error(request, 'Please correct the errors in the form')

    else:
        # GET - show form with current data
        form = LeadEditForm(
            instance=lead,
            company=request.user.company
        )

    context = {
        'form': form,
        'lead': lead,
        'form_title': f'Edit Lead: {lead.name}',
        'submit_text': 'Save Changes',
        'cancel_url': 'leads:lead_detail',
        'cancel_url_kwargs': {'pk': lead.pk},
    }

    return render(request, 'leads/lead_form.html', context)


@login_required
@company_required
@admin_required
@require_POST
def lead_delete_view(request, pk):
    lead = get_object_or_404(
        Lead,
        pk=pk,
        company=request.user.company
    )

    if lead.status == 'won':
        messages.error(
            request,
            'Cannot delete won leads'
        )
        return redirect('leads:lead_detail', pk=lead.pk)

    try:
        lead_name = lead.name

        # Soft delete: Change status to 'deleted'
        lead.status = 'deleted'
        lead.save()

        # Or hard delete (not recommended):
        # lead.delete()

        Activity.objects.create(
            lead=lead,
            user=request.user,
            activity_type='status_changed',
            description=f'Lead deleted by {request.user.get_full_name()}'
        )

        messages.success(
            request,
            f'Lead "{lead_name}" deleted successfully'
        )

        return redirect('leads:lead_list')

    except Exception as e:
        messages.error(
            request,
            f'Error deleting lead: {str(e)}'
        )
        return redirect('leads:lead_detail', pk=lead.pk)


@login_required
@company_required
def lead_assign_view(request, pk):
    lead = get_object_or_404(
        Lead,
        pk=pk,
        company=request.user.company
    )

    if not lead.can_be_assigned():
        messages.error(request, 'Cannot reassign closed leads')
        return redirect('leads:lead_detail', pk=lead.pk)

    if request.method == 'POST':
        form = LeadAssignForm(request.POST, company=request.user.company)

        if form.is_valid():
            new_agent = form.cleaned_data['assigned_to']
            success = lead.assign_to(
                user=new_agent,
                assigned_by=request.user
            )

            if success:
                messages.success(
                    request,
                    f'Lead assigned to {new_agent.get_full_name()}'
                )
            else:
                messages.error(request, 'Assignment failed')

            # Check if AJAX request
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({'success': success})

            return redirect('leads:lead_detail', pk=lead.pk)

    else:
        # GET - show form
        form = LeadAssignForm(
            company=request.user.company,
            initial={'assigned_to': lead.assigned_to}
        )

    context = {
        'form': form,
        'lead': lead,
    }

    return render(request, 'leads/lead_assign.html', context)


@login_required
@company_required
@require_POST
def lead_change_status_view(request, pk):
    lead = get_object_or_404(
        Lead,
        pk=pk,
        company=request.user.company
    )
    new_status = request.POST.get('status')

    if new_status and new_status in dict(Lead.STATUS_CHOICES):
        lead.change_status(new_status, user=request.user)

        messages.success(request,f'Status changed to "{lead.get_status_display()}"')

        # AJAX response
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'status': lead.status,
                'status_display': lead.get_status_display()
            })

        return redirect('leads:lead_detail', pk=lead.pk)

    else:
        messages.error(request, 'Invalid status')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Invalid status'})

        return redirect('leads:lead_detail', pk=lead.pk)


@login_required
@company_required
@require_POST
def lead_change_stage_view(request, pk):
    lead = get_object_or_404(
        Lead,
        pk=pk,
        company=request.user.company
    )

    stage_id = request.POST.get('stage')

    try:
        new_stage = LeadStage.objects.get(pk=stage_id, is_active=True)
        lead.change_stage(new_stage, user=request.user)

        messages.success(request,f'Lead moved to stage "{new_stage.name}"')

        # AJAX response
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'stage': new_stage.id,
                'stage_name': new_stage.name,
                'stage_color': new_stage.color
            })

        return redirect('leads:lead_detail', pk=lead.pk)

    except LeadStage.DoesNotExist:
        messages.error(request, 'Invalid stage')

        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': False, 'error': 'Invalid stage'})

        return redirect('leads:lead_detail', pk=lead.pk)


@login_required
@company_required
@require_POST
def lead_set_follow_up_view(request, pk):
    lead = get_object_or_404(
        Lead,
        pk=pk,
        company=request.user.company
    )

    follow_up_str = request.POST.get('next_follow_up')

    if follow_up_str:
        try:
            # Parse datetime string
            follow_up_dt = datetime.fromisoformat(follow_up_str)

            # Make timezone aware
            follow_up_dt = timezone.make_aware(follow_up_dt)

            # Validate (cannot be in past)
            if follow_up_dt < timezone.now():
                messages.error(request, 'Follow-up date cannot be in the past')
            else:
                # Update lead
                lead.next_follow_up = follow_up_dt
                lead.save()
                Activity.objects.create(
                    lead=lead,
                    user=request.user,
                    activity_type='contacted',
                    description=f'Follow-up scheduled: {follow_up_dt.strftime("%Y-%m-%d %H:%M")}'
                )

                messages.success(request, 'Follow-up date set successfully')

            return redirect('leads:lead_detail', pk=lead.pk)

        except Exception as e:
            messages.error(request, f'Invalid date: {str(e)}')
            return redirect('leads:lead_detail', pk=lead.pk)

    else:
        messages.error(request, 'Please specify a follow-up date')
        return redirect('leads:lead_detail', pk=lead.pk)


@login_required
@company_required
@require_POST
def lead_add_note_view(request, pk):
    lead = get_object_or_404(
        Lead,
        pk=pk,
        company=request.user.company
    )

    form = NoteForm(request.POST)

    if form.is_valid():
        note = lead.add_note(
            content=form.cleaned_data['content'],
            user=request.user
        )

        # AJAX response
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'note': {
                    'id': note.id,
                    'content': note.content,
                    'user': note.user.get_full_name(),
                    'created_at': note.created_at.isoformat(),
                }
            })

        messages.success(request, 'Note added successfully')
        return redirect('leads:lead_detail', pk=lead.pk)

    else:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'errors': form.errors
            })

        messages.error(request, 'Please enter note text')
        return redirect('leads:lead_detail', pk=lead.pk)


@login_required
@company_required
@require_POST
def note_delete_view(request, note_id):
    note = get_object_or_404(
        Note.objects.select_related('lead', 'user'),
        pk=note_id,
        lead__company=request.user.company
    )

    # Permission check
    if not (request.user.is_admin() or note.user == request.user):
        messages.error(request, 'You do not have permission to delete this note')
        return redirect('leads:lead_detail', pk=note.lead.pk)

    lead_pk = note.lead.pk
    note.delete()
    Activity.objects.create(
        lead_id=lead_pk,
        user=request.user,
        activity_type='note_added',
        description=f'Note deleted'
    )

    messages.success(request, 'Note deleted successfully')

    return redirect('leads:lead_detail', pk=lead_pk)


@login_required
@company_required
@require_POST
def lead_quick_update_view(request, pk):
    lead = get_object_or_404(
        Lead,
        pk=pk,
        company=request.user.company
    )
    
    # Permission check - user must be admin or assigned to this lead
    if not (request.user.is_admin() or lead.assigned_to == request.user):
        return JsonResponse({
            'success': False,
            'error': 'You do not have permission to edit this lead'
        }, status=403)

    try:
        # Parse JSON body with error handling
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid JSON format'
            }, status=400)
            
        field = data.get('field')
        value = data.get('value')
        
        # Validate field and value presence
        if not field or value is None:
            return JsonResponse({
                'success': False,
                'error': 'Field and value are required'
            }, status=400)
        
        allowed_fields = ['priority', 'status', 'next_follow_up']

        if field not in allowed_fields:
            return JsonResponse({
                'success': False,
                'error': 'Field not allowed'
            }, status=400)
        
        # Store old value for activity log
        old_value = getattr(lead, field)
        
        # Handle different fields with specific validation
        if field == 'status':
            # Validate status value
            if value not in dict(Lead.STATUS_CHOICES):
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid status value'
                }, status=400)
            # Use model method for status changes
            lead.change_status(value, user=request.user)
            activity_type = 'status_changed'
            
        elif field == 'priority':
            # Validate priority value
            if value not in dict(Lead.PRIORITY_CHOICES):
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid priority value'
                }, status=400)
            setattr(lead, field, value)
            lead.save(update_fields=[field])
            activity_type = 'field_updated'
            
        elif field == 'next_follow_up':
            # Parse and validate datetime
            if value:
                try:
                    follow_up_dt = datetime.fromisoformat(value)
                    follow_up_dt = timezone.make_aware(follow_up_dt)
                    
                    # Validate that date is not in the past
                    if follow_up_dt < timezone.now():
                        return JsonResponse({
                            'success': False,
                            'error': 'Date cannot be in the past'
                        }, status=400)
                    
                    value = follow_up_dt
                except (ValueError, TypeError) as e:
                    return JsonResponse({
                        'success': False,
                        'error': f'Invalid date: {str(e)}'
                    }, status=400)
            
            setattr(lead, field, value)
            lead.save(update_fields=[field])
            activity_type = 'field_updated'
        
        # Create activity with improved description showing old and new values
        old_display = getattr(lead, f'get_{field}_display', lambda: old_value)() if old_value else 'empty'
        new_display = getattr(lead, f'get_{field}_display', lambda: value)()
        
        Activity.objects.create(
            lead=lead,
            user=request.user,
            activity_type=activity_type,
            description=f'Updated {field} from "{old_display}" to "{new_display}"'
        )
        
        display_value = getattr(lead, f'get_{field}_display', lambda: value)()

        return JsonResponse({
            'success': True,
            'field': field,
            'value': str(value) if not isinstance(value, str) else value,
            'display_value': display_value
        })

    except Exception as e:
        # Log the error for debugging
        import logging
        logger = logging.getLogger(__name__)
        logger.error(f"Error in quick update for lead {pk}: {str(e)}")
        
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)




@login_required
@company_required
def lead_kanban_view(request):
    company = request.user.company
    stages = LeadStage.objects.filter(
        is_active=True
    ).order_by('order')

    leads_queryset = Lead.objects.filter(
        company=company).exclude(status='deleted').select_related('source','stage','assigned_to')


    search_query = request.GET.get('search', '').strip()
    if search_query:
        leads_queryset = leads_queryset.filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    filter_form = LeadFilterForm(request.GET, company=company)

    if filter_form.is_valid():
        if filter_form.cleaned_data.get('source'):
            leads_queryset = leads_queryset.filter(
                source=filter_form.cleaned_data['source']
            )

        if filter_form.cleaned_data.get('status'):
            leads_queryset = leads_queryset.filter(
                status=filter_form.cleaned_data['status']
            )

        if filter_form.cleaned_data.get('priority'):
            leads_queryset = leads_queryset.filter(
                priority=filter_form.cleaned_data['priority']
            )

        if filter_form.cleaned_data.get('assigned_to'):
            leads_queryset = leads_queryset.filter(
                assigned_to=filter_form.cleaned_data['assigned_to']
            )

        if filter_form.cleaned_data.get('date_from'):
            leads_queryset = leads_queryset.filter(
                created_at__date__gte=filter_form.cleaned_data['date_from']
            )

        if filter_form.cleaned_data.get('date_to'):
            leads_queryset = leads_queryset.filter(
                created_at__date__lte=filter_form.cleaned_data['date_to']
            )


    stages_data = []
    total_count = 0

    for stage in stages:
        stage_leads = leads_queryset.filter(stage=stage)
        count = stage_leads.count()
        total_count += count

        stages_data.append({
            'stage': stage,
            'leads': stage_leads,
            'count': count,
        })


    context = {
        'stages_data': stages_data,
        'filter_form': filter_form,
        'total_count': total_count,
        'search_query': search_query,
    }

    return render(request, 'leads/lead_kanban.html', context)


@login_required
@company_required
@require_POST
def lead_bulk_actions_view(request):

    try:
        # Parse request data
        if request.content_type == 'application/json':
            data = json.loads(request.body)
        else:
            data = request.POST

        action = data.get('action')
        lead_ids = data.get('lead_ids', '')

        # Parse lead_ids (can be string "1,2,3" or array [1,2,3])
        if isinstance(lead_ids, str):
            lead_ids = [int(id.strip()) for id in lead_ids.split(',') if id.strip()]
        elif isinstance(lead_ids, list):
            lead_ids = [int(id) for id in lead_ids]

        leads = Lead.objects.filter(
            pk__in=lead_ids,
            company=request.user.company
        )

        count = leads.count()

        if count == 0:
            return JsonResponse({
                'success': False,
                'error': 'No leads selected'
            }, status=400)


        if action == 'assign':
            user_id = data.get('user_id')

            try:
                user = User.objects.get(
                    pk=user_id,
                    company=request.user.company
                )

                for lead in leads:
                    lead.assign_to(user, assigned_by=request.user)

                message = f'{count} lead(s) assigned to {user.get_full_name()}'

            except User.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'User not found'
                }, status=404)

        elif action == 'change_status':
            status = data.get('status')

            if status not in dict(Lead.STATUS_CHOICES):
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid status'
                }, status=400)

            for lead in leads:
                lead.change_status(status, user=request.user)

            status_display = dict(Lead.STATUS_CHOICES)[status]
            message = f'{count} lead(s) status changed to "{status_display}"'

        elif action == 'change_stage':
            stage_id = data.get('stage_id')

            try:
                stage = LeadStage.objects.get(pk=stage_id, is_active=True)

                for lead in leads:
                    lead.change_stage(stage, user=request.user)

                message = f'{count} lead(s) moved to stage "{stage.name}"'

            except LeadStage.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': 'Stage not found'
                }, status=404)

        elif action == 'set_priority':
            priority = data.get('priority')

            if priority not in dict(Lead.PRIORITY_CHOICES):
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid priority'
                }, status=400)

            leads.update(priority=priority)

            for lead in leads:
                Activity.objects.create(
                    lead=lead,
                    user=request.user,
                    activity_type='status_changed',
                    description=f'Priority changed to "{dict(Lead.PRIORITY_CHOICES)[priority]}"'
                )

            priority_display = dict(Lead.PRIORITY_CHOICES)[priority]
            message = f'Priority "{priority_display}" set for {count} lead(s)'

        elif action == 'delete':
            # Delete leads (soft delete)
            if not request.user.is_admin():
                return JsonResponse({
                    'success': False,
                    'error': 'Permission denied'
                }, status=403)

            # Soft delete
            leads.update(status='deleted')

            # Log activities
            for lead in leads:
                Activity.objects.create(
                    lead=lead,
                    user=request.user,
                    activity_type='status_changed',
                    description=f'Lead deleted (bulk action)'
                )

            message = f'{count} lead(s) deleted'

        else:
            return JsonResponse({
                'success': False,
                'error': 'Invalid action'
            }, status=400)



        messages.success(request, message)

        # AJAX response
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'count': count,
                'message': message
            })

        return redirect('leads:lead_list')

    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=500)

        messages.error(request, f'Error occurred: {str(e)}')
        return redirect('leads:lead_list')


@login_required
@company_required
def lead_export_view(request):
    export_format = request.GET.get('format', 'excel')
    company = request.user.company
    leads = Lead.objects.filter(company=company).select_related('source', 'stage', 'assigned_to')

    search_query = request.GET.get('search', '').strip()
    if search_query:
        leads = leads.filter(
            Q(name__icontains=search_query) |
            Q(phone__icontains=search_query) |
            Q(email__icontains=search_query)
        )

    filter_form = LeadFilterForm(request.GET, company=company)
    if filter_form.is_valid():
        if filter_form.cleaned_data.get('source'):
            leads = leads.filter(source=filter_form.cleaned_data['source'])
        if filter_form.cleaned_data.get('stage'):
            leads = leads.filter(stage=filter_form.cleaned_data['stage'])
        if filter_form.cleaned_data.get('status'):
            leads = leads.filter(status=filter_form.cleaned_data['status'])
        if filter_form.cleaned_data.get('priority'):
            leads = leads.filter(priority=filter_form.cleaned_data['priority'])
        if filter_form.cleaned_data.get('assigned_to'):
            leads = leads.filter(assigned_to=filter_form.cleaned_data['assigned_to'])


    if export_format == 'excel':
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"

        headers = [
            'ID', 'Name', 'Phone', 'Email',
            'Source', 'Stage', 'Status', 'Priority',
            'Assigned To', 'Created Date', 'Next Follow-up'
        ]

        # Write headers with styling
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")

        # Write data
        for row, lead in enumerate(leads, start=2):
            ws.cell(row=row, column=1, value=lead.id)
            ws.cell(row=row, column=2, value=lead.name)
            ws.cell(row=row, column=3, value=lead.phone)
            ws.cell(row=row, column=4, value=lead.email or '')
            ws.cell(row=row, column=5, value=lead.source.name)
            ws.cell(row=row, column=6, value=lead.stage.name)
            ws.cell(row=row, column=7, value=lead.get_status_display())
            ws.cell(row=row, column=8, value=lead.get_priority_display())
            ws.cell(row=row, column=9, value=lead.assigned_to.get_full_name() if lead.assigned_to else '')
            ws.cell(row=row, column=10, value=lead.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=11,
                    value=lead.next_follow_up.strftime('%Y-%m-%d %H:%M') if lead.next_follow_up else '')

        # Adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Prepare response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response[
            'Content-Disposition'] = f'attachment; filename="leads_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'

        # Save workbook to response
        wb.save(response)

        return response


    elif export_format == 'csv':
        # Create CSV
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="leads_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'

        # Write BOM for Excel UTF-8 compatibility
        response.write('\ufeff')

        writer = csv.writer(response)

        writer.writerow([
            'ID', 'Name', 'Phone', 'Email',
            'Source', 'Stage', 'Status', 'Priority',
            'Assigned To', 'Created Date', 'Next Follow-up'
        ])

        # Data
        for lead in leads:
            writer.writerow([
                lead.id,
                lead.name,
                lead.phone,
                lead.email or '',
                lead.source.name,
                lead.stage.name,
                lead.get_status_display(),
                lead.get_priority_display(),
                lead.assigned_to.get_full_name() if lead.assigned_to else '',
                lead.created_at.strftime('%Y-%m-%d %H:%M'),
                lead.next_follow_up.strftime('%Y-%m-%d %H:%M') if lead.next_follow_up else ''
            ])

        return response

    else:
        messages.error(request, 'Invalid export format')
        return redirect('leads:lead_list')


@login_required
@company_required
@admin_required
def lead_import_view(request):
    if request.method == 'POST':
        form = LeadImportForm(request.POST, request.FILES, company=request.user.company)

        if form.is_valid():
            uploaded_file = form.cleaned_data['file']
            source = form.cleaned_data['source']
            assigned_to = form.cleaned_data.get('assigned_to')

            results = {
                'success': 0,
                'failed': 0,
                'skipped': 0,
                'errors': []
            }

            try:
                file_name = uploaded_file.name.lower()
                rows = []

                if file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                    # Excel file
                    wb = openpyxl.load_workbook(uploaded_file)
                    ws = wb.active

                    # Skip header row, read data rows
                    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                        if row[0]:  # If name exists
                            rows.append({
                                'row_num': row_num,
                                'name': str(row[0]).strip(),
                                'phone': str(row[1]).strip() if row[1] else '',
                                'email': str(row[2]).strip() if len(row) > 2 and row[2] else '',
                                'notes': str(row[3]).strip() if len(row) > 3 and row[3] else ''
                            })

                elif file_name.endswith('.csv'):
                    file_data = TextIOWrapper(uploaded_file.file, encoding='utf-8-sig')
                    csv_reader = csv.reader(file_data)

                    # Skip header
                    next(csv_reader, None)

                    for row_num, row in enumerate(csv_reader, start=2):
                        if row and row[0]:
                            rows.append({
                                'row_num': row_num,
                                'name': row[0].strip(),
                                'phone': row[1].strip() if len(row) > 1 else '',
                                'email': row[2].strip() if len(row) > 2 else '',
                                'notes': row[3].strip() if len(row) > 3 else ''
                            })

                # Get default stage (first Lead stage)
                default_stage = LeadStage.objects.filter(
                    stage_type='lead',
                    is_active=True
                ).order_by('order').first()

                # Get existing phones (to check duplicates)
                existing_phones = set(
                    Lead.objects.filter(
                        company=request.user.company
                    ).values_list('phone', flat=True)
                )

                for row_data in rows:
                    row_num = row_data['row_num']

                    try:
                        if not row_data['name']:
                            results['errors'].append(f"Row {row_num}: Name is required")
                            results['failed'] += 1
                            continue

                        if not row_data['phone']:
                            results['errors'].append(f"Row {row_num}: Phone number is required")
                            results['failed'] += 1
                            continue


                        if row_data['phone'] in existing_phones:
                            results['errors'].append(f"Row {row_num}: Phone number already exists ({row_data['phone']})")
                            results['skipped'] += 1
                            continue


                        lead = Lead.objects.create(
                            company=request.user.company,
                            name=row_data['name'],
                            phone=row_data['phone'],
                            email=row_data['email'] or None,
                            notes=row_data['notes'],
                            source=source,
                            stage=default_stage,
                            status='new',
                            priority='medium',
                            assigned_to=assigned_to
                        )


                        existing_phones.add(row_data['phone'])

                        Activity.objects.create(
                            lead=lead,
                            user=request.user,
                            activity_type='created',
                            description=f'Lead imported from file'
                        )

                        results['success'] += 1

                    except Exception as e:
                        results['errors'].append(f"Row {row_num}: {str(e)}")
                        results['failed'] += 1



                messages.success(
                    request,
                    f'Import completed successfully: {results["success"]} lead(s), '
                    f'skipped: {results["skipped"]}, '
                    f'failed: {results["failed"]}'
                )

                # Store results in session for display
                request.session['import_results'] = results

                return redirect('leads:lead_list')

            except Exception as e:
                messages.error(request, f'Error during import: {str(e)}')

        else:
            messages.error(request, 'Please correct the errors in the form')

    else:
        # GET - show form
        form = LeadImportForm(company=request.user.company)

    context = {
        'form': form,
    }

    return render(request, 'leads/lead_import.html', context)